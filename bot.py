import asyncio
import re
import io
import datetime
import time
import json
import requests
from typing import Dict, Any, Optional, List
from vosk import Model, KaldiRecognizer
import wave
# -------------------------------------------
# НАСТРОЙКИ
# -------------------------------------------
BOT_TOKEN = "8514888342:AAGYavxKcgOaEmtHFSydpFze3x9Uw_bh5SE"
ADMIN_ID = 1750883753
PAGE_SIZE = 5  # сколько товаров показывать на странице в каталоге

# источник базы товаров (GitHub RAW)
PRODUCTS_URL = "https://raw.githubusercontent.com/jon199835-crypto/mar_shopping_bot/main/products.json"

# КЭШ JSON-файла из GitHub
DB_CACHE: List[Dict[str, Any]] = []
DB_LAST_UPDATE = 0  # timestamp последнего обновления кэша

# -------------------------------------------
# AIoGram
# -------------------------------------------
from aiogram import Bot, Dispatcher, F
from aiogram.types import (
    Message,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
    BufferedInputFile,
    ReplyKeyboardMarkup,
    KeyboardButton,
)
from aiogram.filters import Command

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

from openpyxl import load_workbook

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
    Table,
    TableStyle,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# -------------------------------------------
# ХРАНЕНИЕ ПОЛЬЗОВАТЕЛЕЙ / СОСТОЯНИЙ
# -------------------------------------------

# user_id -> { article -> { name, price_opt(int), qty(int) } }
USER_CARTS: Dict[int, Dict[str, Dict[str, Any]]] = {}

# временное хранение ввода количества через numpad:
# user_id -> {"article": str, "qty": str}
QTY_INPUT: Dict[int, Dict[str, str]] = {}

# article -> file_id (фото в телеге, чтобы слать мгновенно)
PHOTO_CACHE: Dict[str, str] = {}

# user_id — кто уже видел приветствие
FIRST_VISIT = set()

# -------------------------------------------
# КЛАВИАТУРЫ
# -------------------------------------------

MAIN_MENU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔎 Найти артикул")],
        [KeyboardButton(text="🧺 Корзина"), KeyboardButton(text="📄 Оформить заказ")],
        [KeyboardButton(text="📚 Инструкция"), KeyboardButton(text="📞 Контакты")],
        [KeyboardButton(text="📂 Каталог моделей")],
        [KeyboardButton(text="📤 Загрузить Excel")],
    ],
    resize_keyboard=True,
)

# ЧИСЛОВАЯ КЛАВИАТУРА (NUMPAD)
NUMPAD = InlineKeyboardMarkup(
    inline_keyboard=[
        [
            InlineKeyboardButton(text="1", callback_data="qty_digit_1"),
            InlineKeyboardButton(text="2", callback_data="qty_digit_2"),
            InlineKeyboardButton(text="3", callback_data="qty_digit_3"),
        ],
        [
            InlineKeyboardButton(text="4", callback_data="qty_digit_4"),
            InlineKeyboardButton(text="5", callback_data="qty_digit_5"),
            InlineKeyboardButton(text="6", callback_data="qty_digit_6"),
        ],
        [
            InlineKeyboardButton(text="7", callback_data="qty_digit_7"),
            InlineKeyboardButton(text="8", callback_data="qty_digit_8"),
            InlineKeyboardButton(text="9", callback_data="qty_digit_9"),
        ],
        [
            InlineKeyboardButton(text="0", callback_data="qty_digit_0"),
            InlineKeyboardButton(text="⌫", callback_data="qty_digit_back"),
            InlineKeyboardButton(text="✔️ OK", callback_data="qty_digit_ok"),
        ],
    ]
)

# -------------------------------------------
# ЗАГРУЗКА JSON С GitHub
# -------------------------------------------
def recognize_speech_vosk(wav_bytes: bytes) -> str:
    """
    Стабильное распознавание через Vosk:
    - тихий лог
    - русский язык
    - стабильный sample rate 16000
    """
    from vosk import Model, KaldiRecognizer, SetLogLevel
    SetLogLevel(-1)

    # Загружаем WAV
    wf = wave.open(io.BytesIO(wav_bytes), "rb")

    # Проверяем sample rate (должен быть 16000)
    rate = wf.getframerate()
    if rate != 16000:
        print(f"[WARN] WAV sample rate = {rate}, ожидалось 16000!")

    # Модель берём из папки "model"
    model = Model("model")
    rec = KaldiRecognizer(model, 16000)

    text = ""

    while True:
        data = wf.readframes(4000)
        if len(data) == 0:
            break

        if rec.AcceptWaveform(data):
            chunk = json.loads(rec.Result()).get("text", "")
            if chunk:
                text += chunk + " "

    final = json.loads(rec.FinalResult()).get("text", "")
    text += final

    return text.strip()
def load_db() -> List[Dict[str, Any]]:
    """
    Кэшируем products.json на 60 секунд.
    Формат записи:
    {
      "article": "08-4300",
      "name": "...",
      "opt_price": "3058",
      "rrc_price": "9627",
      "photo_url": "https://...",
      "stock": 62,
      "model": "Yamaha Viking 540"
    }
    """
    global DB_CACHE, DB_LAST_UPDATE

    now = time.time()
    if now - DB_LAST_UPDATE > 60 or not DB_CACHE:
        try:
            resp = requests.get(PRODUCTS_URL, timeout=7)
            resp.raise_for_status()
            DB_CACHE = json.loads(resp.text)
            DB_LAST_UPDATE = now
            print(f"[DB] Обновлена, всего записей: {len(DB_CACHE)}")
        except Exception as e:
            print("[DB] Ошибка загрузки JSON:", e)

    return DB_CACHE

def search_products_by_name(query: str) -> List[Dict[str, Any]]:
    """
    Ищет товары по части названия (регистр неважен).
    """
    db = load_db()
    q = query.lower().strip()

    results = []
    for p in db:
        name = str(p.get("name", "")).lower()
        article = str(p.get("article", "")).lower()

        # игнорируем модели и мусор
        if not name:
            continue

        # ищем по вхождению
        if q in name:
            results.append(p)

    return results
    
def get_product_by_article(article_query: str) -> Optional[Dict[str, Any]]:
    """
    Находим товар по артикулу (регистронезависимо, без лишних пробелов).
    """
    db = load_db()
    query = article_query.strip().lower()

    for p in db:
        article = str(p.get("article", "")).strip().lower()
        if article == query:
            return p

    return None


def get_products_by_model(model_name: str) -> List[Dict[str, Any]]:
    """
    Возвращает товары по точному названию модели (регистр не важен).
    """
    db = load_db()
    m = model_name.strip().lower()
    return [p for p in db if str(p.get("model", "")).strip().lower() == m]


def get_all_models() -> List[str]:
    """
    Список всех моделей (без пустых, уникальный, отсортированный).
    """
    db = load_db()
    models = set()

    for p in db:
        model_val = str(p.get("model", "")).strip()
        if model_val:
            models.add(model_val)

    return sorted(models)


# -------------------------------------------
# ПОМОЩНИКИ
# -------------------------------------------

def parse_price_to_int(val: Any) -> int:
    s = str(val).replace(" ", "").replace("\xa0", "")
    return int(s) if s.isdigit() else 0


def resolve_real_url(url: str) -> str:
    try:
        r = requests.get(url, allow_redirects=True, timeout=7)
        return r.url
    except Exception:
        return url


def parse_article_and_qty(text: str):
    """
    Поддерживаем форматы:
    - '8512-153-19'
    - '8512-153-19 x 3' / '8512-153-19 х 3'
    - '8512-153-19 * 5'
    - '8512-153-19 10'
    """
    s = text.strip()
    low = s.lower().replace("х", "x")

    m = re.match(r"^(.+?)\s*[x\*]\s*(\d+)$", low)
    if m:
        return m.group(1).strip(), int(m.group(2))

    m2 = re.match(r"^(.+)\s+(\d+)$", s)
    if m2:
        return m2.group(1).strip(), int(m2.group(2))

    return s, None


# -------------------------------------------
# ОТОБРАЖЕНИЕ ТОВАРА
# -------------------------------------------

async def send_product_card(message: Message, product: Dict[str, Any]) -> None:
    article = str(product.get("article", "")).strip()
    name = str(product.get("name", article))
    opt_price_str = str(product.get("opt_price", "0"))
    photo_url = str(product.get("photo_url", "")).strip()

    stock_raw = product.get("stock", 0)
    try:
        stock = int(stock_raw)
    except Exception:
        stock = 0

    caption = (
        f"📦 *{name}*\n"
        f"🆔 Артикул: `{article}`\n\n"
        f"📦 Наличие: *{stock} шт*\n\n"
        f"💰 Опт: *{opt_price_str} ₽*"
    )

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="➕1", callback_data=f"add_1_{article}"),
                InlineKeyboardButton(text="➕2", callback_data=f"add_2_{article}"),
                InlineKeyboardButton(text="➕5", callback_data=f"add_5_{article}"),
                InlineKeyboardButton(text="➕10", callback_data=f"add_10_{article}"),
            ],
            [
                InlineKeyboardButton(
                    text="✏️ Ввести количество",
                    callback_data=f"add_manual_{article}",
                )
            ],
            [
                InlineKeyboardButton(
                    text="🧺 Открыть корзину",
                    callback_data="open_cart",
                )
            ],
        ]
    )

    # фото из кэша
    if article in PHOTO_CACHE:
        file_id = PHOTO_CACHE[article]
        try:
            await message.answer_document(
                file_id,
                caption=caption,
                parse_mode="Markdown",
                reply_markup=kb,
            )
            return
        except Exception:
            del PHOTO_CACHE[article]

    # Качаем фото
    if photo_url.startswith("http"):
        real_url = resolve_real_url(photo_url)
        try:
            resp = requests.get(real_url, timeout=7)
            resp.raise_for_status()
            img_bytes = io.BytesIO(resp.content)
        except Exception:
            await message.answer(caption, parse_mode="Markdown", reply_markup=kb)
            return

        thumb_bytes = None
        if PILImage is not None:
            try:
                im = PILImage.open(img_bytes)
                im.thumbnail((200, 120))
                thumb_io = io.BytesIO()
                im.save(thumb_io, format="JPEG")
                thumb_io.seek(0)
                thumb_bytes = thumb_io.getvalue()
            except Exception:
                thumb_bytes = None

        img_bytes.seek(0)

        sent = await message.answer_document(
            document=BufferedInputFile(
                img_bytes.getvalue(), filename=f"{article}.jpg"
            ),
            thumb=(
                BufferedInputFile(
                    thumb_bytes, filename=f"{article}_thumb.jpg"
                )
                if thumb_bytes
                else None
            ),
            caption=caption,
            parse_mode="Markdown",
            reply_markup=kb,
        )

        if sent.document:
            PHOTO_CACHE[article] = sent.document.file_id
        return

    await message.answer(caption, parse_mode="Markdown", reply_markup=kb)


# -------------------------------------------
# КОРЗИНА
# -------------------------------------------

async def send_cart(message_or_cb_msg: Message, user_id: int, edit: bool = False) -> None:
    cart = USER_CARTS.get(user_id, {})

    if not cart:
        await message_or_cb_msg.answer("🧺 Корзина пуста.")
        return

    if edit:
        try:
            await message_or_cb_msg.delete()
        except Exception:
            pass

    total = 0

    for article, item in cart.items():
        qty = item["qty"]
        price = item["price_opt"]
        name = item["name"]
        subtotal = qty * price
        total += subtotal

        text = (
            f"🔹 *{name}*\n"
            f"🆔 `{article}`\n"
            f"Кол-во: *{qty}* × {price} ₽ = *{subtotal} ₽*"
        )

        kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="➖", callback_data=f"cart_minus_{article}"
                    ),
                    InlineKeyboardButton(
                        text="➕", callback_data=f"cart_plus_{article}"
                    ),
                ]
            ]
        )

        await message_or_cb_msg.answer(text, parse_mode="Markdown", reply_markup=kb)

    total_text = f"💰 *Итого: {total} ₽*"

    kb_total = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="🧹 Очистить корзину", callback_data="cart_clear"
                )
            ],
            [
                InlineKeyboardButton(
                    text="📄 Оформить заказ", callback_data="checkout"
                )
            ],
        ]
    )

    await message_or_cb_msg.answer(
        total_text, parse_mode="Markdown", reply_markup=kb_total
    )


def add_to_cart(user_id: int, product: Dict[str, Any], qty: int) -> bool:
    if qty <= 0:
        return False

    stock_raw = product.get("stock", 0)
    try:
        stock = int(stock_raw)
    except Exception:
        stock = 0

    article = str(product.get("article", "")).strip()

    if user_id not in USER_CARTS:
        USER_CARTS[user_id] = {}

    current_qty = USER_CARTS[user_id].get(article, {}).get("qty", 0)

    if current_qty + qty > stock:
        return False

    name = str(product.get("name", article))
    opt_price_int = parse_price_to_int(product.get("opt_price", "0"))

    if article not in USER_CARTS[user_id]:
        USER_CARTS[user_id][article] = {
            "name": name,
            "price_opt": opt_price_int,
            "qty": 0,
        }

    USER_CARTS[user_id][article]["qty"] += qty
    return True


def change_cart_qty(user_id: int, article: str, delta: int) -> None:
    if user_id not in USER_CARTS:
        return
    if article not in USER_CARTS[user_id]:
        return

    USER_CARTS[user_id][article]["qty"] += delta
    if USER_CARTS[user_id][article]["qty"] <= 0:
        del USER_CARTS[user_id][article]


# -------------------------------------------
# КАТАЛОГ МОДЕЛЕЙ
# -------------------------------------------

async def send_model_page(message: Message, model: str, page: int):
    products = get_products_by_model(model)
    if not products:
        await message.answer("❌ Для этой модели запчастей не найдено.")
        return

    total = len(products)
    pages = (total + PAGE_SIZE - 1) // PAGE_SIZE

    page = max(1, min(page, pages))

    start = (page - 1) * PAGE_SIZE
    end = start + PAGE_SIZE
    page_products = products[start:end]

    await message.answer(
        f"📂 Запчасти для *{model}* (стр. {page}/{pages}):",
        parse_mode="Markdown",
    )

    for p in page_products:
        await send_product_card(message, p)

    if pages > 1:
        buttons = []
        if page > 1:
            buttons.append(
                InlineKeyboardButton(
                    text="⬅️ Назад", callback_data=f"modelpage_{page-1}_{model}"
                )
            )
        if page < pages:
            buttons.append(
                InlineKeyboardButton(
                    text="➡️ Далее", callback_data=f"modelpage_{page+1}_{model}"
                )
            )

        kb = InlineKeyboardMarkup(inline_keyboard=[buttons])
        await message.answer(f"Страница {page}/{pages}", reply_markup=kb)


# -------------------------------------------
# TELEGRAM BOT
# -------------------------------------------

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


@dp.message(Command("start"))
async def cmd_start(message: Message):
    user_id = message.from_user.id

    if user_id not in FIRST_VISIT:
        FIRST_VISIT.add(user_id)

        await message.answer(
            "👋 Привет! Я бот для заказа запчастей.\n\n"
            "🔎 Чтобы начать — просто отправьте артикул, например:\n"
            "`8512-153-19`\n\n"
            "Или используйте главное меню ниже 👇",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU,
        )
        return

    await message.answer("Вы снова в боте 😊\nВыберите действие:", reply_markup=MAIN_MENU)


# -------------------------------------------
# ГЛАВНОЕ МЕНЮ
# -------------------------------------------

@dp.message(F.text == "🔎 Найти артикул")
async def btn_find_article(message: Message):
    await message.answer(
        "Введите артикул, например:\n`8512-153-19`",
        parse_mode="Markdown",
    )


@dp.message(F.text == "🧺 Корзина")
async def btn_cart(message: Message):
    await send_cart(message, message.from_user.id)


@dp.message(F.text == "📄 Оформить заказ")
async def btn_checkout(message: Message):
    fake_callback = type(
        "obj", (object,), {"from_user": message.from_user, "message": message}
    )
    await checkout_handler(fake_callback)


@dp.message(F.text == "📚 Инструкция")
async def btn_instruction(message: Message):
    await message.answer(
        "📚 *Как пользоваться ботом:*\n\n"
        "1️⃣ Введите артикул\n"
        "2️⃣ Добавьте количество\n"
        "3️⃣ Откройте корзину\n"
        "4️⃣ Нажмите «Оформить заказ»\n\n"
        "Бот сформирует PDF и отправит менеджеру.",
        parse_mode="Markdown",
    )


@dp.message(F.text == "📞 Контакты")
async def btn_contacts(message: Message):
    await message.answer(
        "📞 *Контакты:*\n\n"
        "Менеджер: @evgenijtuzikov\n"
        "Телефон: +7...\n"
        "Работаем ежедневно 10:00–21:00",
        parse_mode="Markdown",
    )


@dp.message(F.text == "📂 Каталог моделей")
async def show_model_catalog(message: Message):
    models = get_all_models()

    if not models:
        await message.answer("❌ В базе нет ни одной модели.")
        return

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=m, callback_data=f"model_{m}")] for m in models
        ]
    )

    await message.answer("Выберите модель снегохода:", reply_markup=kb)


@dp.message(F.text == "📤 Загрузить Excel")
async def btn_upload_excel(message: Message):
    await message.answer(
        "📤 *Загрузка Excel-прайса*\n\n"
        "Отправьте файл формата *.xlsx*, содержащий:\n"
        "`Артикул | Количество`\n\n"
        "Пример:\n"
        "`8512-153-19 | 3`\n"
        "`3B4-23311-00 | 1`\n\n"
        "После загрузки я автоматически добавлю товары в корзину.",
        parse_mode="Markdown",
    )

@dp.message(F.voice)
async def voice_handler(message: Message):
    user_id = message.from_user.id

    # 1. Скачиваем OGG
    voice_file = await bot.download(message.voice.file_id)
    ogg_bytes = voice_file.read()

    # 2. Конвертируем в WAV (жёстко 16000 Hz)
    from pydub import AudioSegment

    audio = AudioSegment.from_file(io.BytesIO(ogg_bytes), format="ogg")
    audio = audio.set_frame_rate(16000).set_channels(1)

    wav_io = io.BytesIO()
    # ffmpeg принудительно выставляет частоту
    audio.export(wav_io, format="wav", parameters=["-ar", "16000"])
    wav_bytes = wav_io.getvalue()

    # 3. ЛОГ: проверим частоту
    try:
        wf_test = wave.open(io.BytesIO(wav_bytes), "rb")
        print(f"[DEBUG] WAV rate = {wf_test.getframerate()} Hz")
    except:
        print("[ERROR] Не удалось открыть экспортированный WAV")

    # 4. Распознаём
    text = recognize_speech_vosk(wav_bytes)

    if not text:
        await message.answer("Не расслышал 🙈 Попробуйте ещё раз.")
        return

    await message.answer(f"🎤 Вы сказали: *{text}*", parse_mode="Markdown")

    # 5. Пытаемся распознать товар / артикул
    article_query, qty = parse_article_and_qty(text)
    product = get_product_by_article(article_query)

    if product:
        return await send_product_card(message, product)

    # 6. Поиск по названию
    results = search_products_by_name(text)

    if not results:
        await message.answer("❌ Ничего не найдено по вашему запросу.")
        return

    if len(results) == 1:
        return await send_product_card(message, results[0])

    await message.answer(
        f"🔎 Найдено {len(results)} позиций, показываю первые 10:",
        parse_mode="Markdown"
    )

    for p in results[:10]:
        await send_product_card(message, p)
# -------------------------------------------
# ОБРАБОТКА EXCEL
# -------------------------------------------

@dp.message(F.document)
async def handle_excel_upload(message: Message):
    user_id = message.from_user.id
    file = message.document

    if not file.file_name.lower().endswith(".xlsx"):
        await message.answer("Пожалуйста, отправьте файл Excel в формате .xlsx")
        return

    file_bytes = await bot.download(file)
    file_bytes.seek(0)

    try:
        wb = load_workbook(file_bytes, data_only=True)
        ws = wb.active
    except Exception:
        await message.answer("Не удалось прочитать Excel-файл 😔")
        return

    added = 0
    errors = []

    header_map = {}
    first_row = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    for idx, title in enumerate(first_row):
        if "артикул" in title:
            header_map["article"] = idx
        if "кол" in title:
            header_map["qty"] = idx

    if not header_map:
        header_map = {"article": 0, "qty": 1}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[header_map["article"]]:
            continue

        raw_article = row[header_map["article"]]

        if isinstance(raw_article, (int, float)):
            raw_article = str(raw_article).rstrip(".0")

        article = str(raw_article).strip()
        qty_raw = row[header_map["qty"]]

        try:
            qty = int(qty_raw)
            if qty <= 0:
                raise ValueError
        except Exception:
            errors.append(f"{article} — неверное количество")
            continue

        product = get_product_by_article(article)
        if not product:
            errors.append(f"{article} — товар не найден")
            continue

        ok = add_to_cart(user_id, product, qty)
        if not ok:
            stock_raw = product.get("stock", 0)
            try:
                stock = int(stock_raw)
            except Exception:
                stock = 0
            errors.append(f"{article} — недостаточно на складе ({stock})")
            continue

        added += 1

    msg = f"📥 Загрузка Excel завершена!\n\n"
    msg += f"✅ Добавлено позиций: *{added}*\n"

    if errors:
        msg += "\n⚠️ Ошибки:\n" + "\n".join(f"• {e}" for e in errors)

    await message.answer(msg, parse_mode="Markdown")

    if added > 0:
        await send_cart(message, user_id)


# -------------------------------------------
# ОБРАБОТЧИК СООБЩЕНИЙ (ПОИСК)
# -------------------------------------------

@dp.message()
async def handle_message(message: Message):
    text = message.text.strip()
    user_id = message.from_user.id

    # --- сначала попытка распознать как артикул ---
    article_query, qty = parse_article_and_qty(text)
    product = get_product_by_article(article_query)

    if product:  # нашли артикул
        if qty:
            ok = add_to_cart(user_id, product, qty)
            if not ok:
                stock = int(product.get("stock", 0))
                await message.answer(f"❗ Доступно только {stock} шт")
                return

            await message.answer(
                f"✅ Добавлено {qty} шт *{product['name']}* (арт. `{product['article']}`)",
                parse_mode="Markdown"
            )
            await send_cart(message, user_id)
            return

        return await send_product_card(message, product)

    # --- если артикул не найден → ищем по названию ---
    results = search_products_by_name(text)

    if not results:
        await message.answer("❌ Ничего не найдено по вашему запросу.")
        return

    # Если одна позиция — показываем карточку
    if len(results) == 1:
        await send_product_card(message, results[0])
        return

    # Если много — выдаём первые 10
    msg = f"🔎 Найдено {len(results)} позиций по запросу: *{text}*\nПоказываю первые 10:"
    await message.answer(msg, parse_mode="Markdown")

    for p in results[:10]:
        await send_product_card(message, p)


# -------------------------------------------
# NUMPAD — ВВОД КОЛИЧЕСТВА
# -------------------------------------------

@dp.callback_query(F.data.startswith("add_manual_"))
async def cb_manual_qty(callback: CallbackQuery):
    """Нажата кнопка '✏️ Ввести количество' — показываем numpad."""
    user_id = callback.from_user.id
    article = callback.data.replace("add_manual_", "")

    QTY_INPUT[user_id] = {"article": article, "qty": ""}

    await callback.message.answer(
        f"Введите количество для `{article}`:\nТекущее: *пусто*",
        reply_markup=NUMPAD,
        parse_mode="Markdown",
    )
    await callback.answer()


@dp.callback_query(F.data.startswith("qty_digit_"))
async def cb_numpad(callback: CallbackQuery):
    user_id = callback.from_user.id

    if user_id not in QTY_INPUT:
        await callback.answer()
        return

    action = callback.data.replace("qty_digit_", "")
    current = QTY_INPUT[user_id]["qty"]
    article = QTY_INPUT[user_id]["article"]

    # --- ЦИФРЫ ---
    if action.isdigit():
        if len(current) < 4:  # ограничение длины
            QTY_INPUT[user_id]["qty"] += action

    # --- СТЕРЕТЬ ---
    elif action == "back":
        QTY_INPUT[user_id]["qty"] = current[:-1]

    # --- OK ---
    elif action == "ok":
        qty_text = QTY_INPUT[user_id]["qty"]

        if qty_text == "":
            await callback.answer("Введите количество!", show_alert=True)
            return

        qty = int(qty_text)
        product = get_product_by_article(article)

        if not product:
            await callback.answer("Товар не найден.", show_alert=True)
            return   # ❗ не удаляем QTY_INPUT — позволяем ввести заново

        # --- ПРОВЕРКА НАЛИЧИЯ ---
        if not add_to_cart(user_id, product, qty):
            stock_raw = product.get("stock", 0)
            try:
                stock = int(stock_raw)
            except:
                stock = 0

            await callback.answer(
                f"Недостаточно на складе! Доступно: {stock}",
                show_alert=True
            )

            return  # ❗ не удаляем QTY_INPUT → numpad остаётся активным

        # --- УСПЕХ ---
        del QTY_INPUT[user_id]

        await callback.message.answer(
            f"✅ Добавлено {qty} шт товара *{product['name']}* (`{article}`)",
            parse_mode="Markdown",
        )
        await send_cart(callback.message, user_id)
        await callback.answer()
        return

    # --- обновляем текст numpad ---
    new_val = QTY_INPUT[user_id]["qty"] or "пусто"

    try:
        await callback.message.edit_text(
            f"Введите количество для `{article}`:\nТекущее: *{new_val}*",
            reply_markup=NUMPAD,
            parse_mode="Markdown",
        )
    except:
        pass

    await callback.answer()


# -------------------------------------------
# CALLBACK: ОТКРЫТЬ / ОЧИСТИТЬ КОРЗИНУ
# -------------------------------------------

@dp.callback_query(F.data == "open_cart")
async def cb_open_cart(callback: CallbackQuery):
    await callback.answer()
    await send_cart(callback.message, callback.from_user.id)


@dp.callback_query(F.data == "cart_clear")
async def cb_cart_clear(callback: CallbackQuery):
    user_id = callback.from_user.id
    USER_CARTS[user_id] = {}
    await callback.answer("Корзина очищена.")
    await callback.message.answer("🧺 Корзина очищена.")


# -------------------------------------------
# CALLBACK: БЫСТРЫЕ КНОПКИ ДОБАВЛЕНИЯ (+1,+2,+5,+10)
# -------------------------------------------

@dp.callback_query(F.data.regexp(r"^add_(\d+)_"))
async def cb_add_quick(callback: CallbackQuery):
    user_id = callback.from_user.id
    data = callback.data  # add_5_ARTICLE

    m = re.match(r"^add_(\d+)_(.+)$", data)
    if not m:
        await callback.answer("Ошибка формата.", show_alert=True)
        return

    qty = int(m.group(1))
    article = m.group(2)

    product = get_product_by_article(article)
    if not product:
        await callback.answer("Товар не найден.", show_alert=True)
        return

    ok = add_to_cart(user_id, product, qty)
    if not ok:
        stock_raw = product.get("stock", 0)
        try:
            stock = int(stock_raw)
        except Exception:
            stock = 0
        await callback.answer(
            f"❗ На складе доступно только {stock} шт", show_alert=True
        )
        return

    await callback.answer(f"Добавлено {qty} шт в корзину!")


# -------------------------------------------
# CALLBACK: КАТАЛОГ МОДЕЛЕЙ
# -------------------------------------------

@dp.callback_query(F.data.startswith("model_"))
async def cb_show_model_parts(callback: CallbackQuery):
    model = callback.data.replace("model_", "", 1)
    await callback.answer()
    await send_model_page(callback.message, model, page=1)


@dp.callback_query(F.data.startswith("modelpage_"))
async def cb_model_page(callback: CallbackQuery):
    data = callback.data
    _, page_str, model = data.split("_", 2)

    try:
        page = int(page_str)
    except ValueError:
        await callback.answer("Ошибка страницы.", show_alert=True)
        return

    try:
        await callback.message.delete()
    except Exception:
        pass

    await callback.answer()
    await send_model_page(callback.message, model, page)


# -------------------------------------------
# CALLBACK: ПЛЮС / МИНУС В КОРЗИНЕ
# -------------------------------------------

@dp.callback_query(F.data.startswith("cart_plus_"))
async def cb_cart_plus(callback: CallbackQuery):
    user_id = callback.from_user.id
    article = callback.data.replace("cart_plus_", "", 1)

    product = get_product_by_article(article)
    if not product:
        await callback.answer("Товар не найден.", show_alert=True)
        return

    ok = add_to_cart(user_id, product, 1)
    if not ok:
        stock_raw = product.get("stock", 0)
        try:
            stock = int(stock_raw)
        except Exception:
            stock = 0
        await callback.answer(
            f"❗ Доступно только {stock} шт", show_alert=True
        )
        return

    await callback.answer("Увеличено.")
    await send_cart(callback.message, user_id, edit=True)


@dp.callback_query(F.data.startswith("cart_minus_"))
async def cb_cart_minus(callback: CallbackQuery):
    user_id = callback.from_user.id
    article = callback.data.replace("cart_minus_", "", 1)

    change_cart_qty(user_id, article, -1)

    await callback.answer("Уменьшено.")
    await send_cart(callback.message, user_id, edit=True)


# -------------------------------------------
# CALLBACK: ОФОРМИТЬ ЗАКАЗ (PDF — СТАРЫЙ ДИЗАЙН)
# -------------------------------------------

@dp.callback_query(F.data == "checkout")
async def checkout_handler(callback: CallbackQuery):
    user_id = callback.from_user.id
    cart = USER_CARTS.get(user_id, {})

    if not cart:
        await callback.answer("Корзина пуста!", show_alert=True)
        return

    pdfmetrics.registerFont(TTFont("DejaVu", "DejaVuSans.ttf"))
    pdfmetrics.registerFont(TTFont("DejaVu-Bold", "DejaVuSans-Bold.ttf"))

    styles = getSampleStyleSheet()
    for s in styles.byName:
        styles[s].fontName = "DejaVu"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Заказ Моторешение")

    elems = []
    elems.append(Paragraph("<b>Заказ Моторешение</b>", styles["Title"]))
    elems.append(Spacer(1, 12))

    elems.append(
        Paragraph(
            f"Дата: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["Normal"],
        )
    )
    user_label = callback.from_user.username or f"id {user_id}"
    elems.append(Paragraph(f"Клиент: @{user_label}", styles["Normal"]))
    elems.append(Spacer(1, 20))

    table_data = [["Фото", "Артикул", "Название", "Кол-во", "Цена", "Сумма"]]

    total_sum = 0

    for article, item in cart.items():
        name = item["name"]
        qty = item["qty"]
        price = item["price_opt"]
        subtotal = qty * price
        total_sum += subtotal

        product = get_product_by_article(article)
        photo_url = product["photo_url"] if product else ""

        if photo_url and str(photo_url).startswith("http"):
            try:
                resp = requests.get(photo_url, timeout=5)
                img_bytes = io.BytesIO(resp.content)
                img_obj = Image(img_bytes, width=50, height=50)
            except Exception:
                img_obj = Paragraph("Нет фото", styles["Normal"])
        else:
            img_obj = Paragraph("Нет фото", styles["Normal"])

        name_paragraph = Paragraph(name, styles["Normal"])

        table_data.append(
            [
                img_obj,
                article,
                name_paragraph,
                Paragraph(f"{qty}", styles["Normal"]),
                Paragraph(f"{price} ₽", styles["Normal"]),
                Paragraph(f"{subtotal} ₽", styles["Normal"]),
            ]
        )

    table = Table(table_data, colWidths=[60, 55, 180, 50, 55, 60])

    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (3, 1), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, -1), "DejaVu"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "DejaVu-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
            ]
        )
    )

    elems.append(table)
    elems.append(Spacer(1, 20))
    elems.append(Paragraph(f"<b>Итого: {total_sum} ₽</b>", styles["Heading2"]))

    doc.build(elems)

    buffer.seek(0)
    pdf_bytes = buffer.getvalue()

    file_for_user = BufferedInputFile(pdf_bytes, filename="Заказ Моторешение.pdf")
    file_for_admin = BufferedInputFile(pdf_bytes, filename="Заказ Моторешение.pdf")

    await callback.message.answer_document(
        document=file_for_user,
        caption="📄 Ваш заказ сформирован!",
    )

    await bot.send_document(
        ADMIN_ID,
        document=file_for_admin,
        caption=(
            "📥 Новый заказ из бота\n"
            f"Клиент: {callback.from_user.full_name}\n"
            f"Username: @{callback.from_user.username}\n"
            f"ID: {callback.from_user.id}"
        ),
    )

    await callback.answer("PDF заказ сформирован!")


# -------------------------------------------
# RUN
# -------------------------------------------

async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
